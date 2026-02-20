import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import os
from typing import List, Dict

class ExcelGenerator:
    """Класс для создания Excel таблиц с бронями"""
    
    @staticmethod
    def get_waiter_name_for_table(table_number: str, date: str, db) -> str:
        """
        Получает имя официанта для конкретного стола на указанную дату
        """
        try:
            waiters = db.get_waiters_for_table_on_date_with_names(table_number, date)
            if waiters:
                # Если несколько официантов на один стол, объединяем имена
                return ", ".join(waiters)
            return ""
        except Exception as e:
            print(f"Ошибка при получении имени официанта: {e}")
            return ""
    
    @staticmethod
    def get_deposit_status_symbol(deposit: int, deposit_paid: int) -> str:
        """
        Возвращает символ статуса депозита: ❌ или ✅
        """
        if deposit > 0:
            return "✅" if deposit_paid == 1 else "❌"
        return ""
    
    @staticmethod
    def create_reservation_file(reservations: List[Dict], date: str, db=None) -> str:
        """
        Создает Excel файл с бронями на указанную дату
        reservations: список словарей с данными броней
        date: дата в формате YYYY-MM-DD
        db: объект базы данных для получения имен официантов
        Возвращает путь к созданному файлу
        """
        # Создаем новую книгу Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Брони {date}"
        
        # Заголовки столбцов (ДОБАВЛЕН СТОЛБЕЦ "Статус депозита")
        headers = [
            "ID", "Дата", "Стол", "Имя гостя", "Телефон", 
            "Повод", "Время", "Гостей", "Депозит (₽)", "Статус депозита", "Официант"
        ]
        
        # Форматирование заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Применяем форматирование к заголовкам
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Заполняем данными
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'),
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        for row_num, res in enumerate(reservations, 2):
            table_number = res.get('table_number', '')
            deposit = res.get('deposit', 0)
            deposit_paid = res.get('deposit_paid', 0)
            
            # Получаем имя официанта для этого стола
            waiter_name = ""
            if db and table_number:
                waiter_name = ExcelGenerator.get_waiter_name_for_table(table_number, date, db)
            
            # Получаем символ статуса депозита
            deposit_status = ExcelGenerator.get_deposit_status_symbol(deposit, deposit_paid)
            
            row_data = [
                res.get('id', ''),                    # ID
                res.get('date', ''),                   # Дата
                table_number if table_number else 'Не назначен',  # Стол
                res.get('name', ''),                    # Имя
                res.get('phone', ''),                   # Телефон
                res.get('occasion', '-'),                # Повод
                res.get('time', ''),                     # Время
                res.get('guests', ''),                   # Гости
                deposit if deposit > 0 else '',          # Депозит
                deposit_status,                          # Статус депозита (❌ или ✅)
                waiter_name                               # Официант
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                
                # Форматирование депозита
                if col_num == 9 and value and int(value) > 0:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif col_num == 9 and value and int(value) == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                # Форматирование статуса депозита
                if col_num == 10 and value == "✅":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif col_num == 10 and value == "❌":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Автоматическая ширина колонок
        for col in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in range(1, len(reservations) + 2):
                cell = ws[f"{column_letter}{row}"]
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)  # Не шире 30 символов
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем файл
        timestamp = datetime.datetime.now().strftime('%H%M%S')
        filename = f"reservations_{date}_{timestamp}.xlsx"
        filepath = os.path.join("excel_files", filename)
        
        # Создаем папку, если её нет
        os.makedirs("excel_files", exist_ok=True)
        
        wb.save(filepath)
        print(f"✅ Excel файл сохранен: {filepath}")
        return filepath